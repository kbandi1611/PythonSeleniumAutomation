import openpyxl

# to write same data in all the cells
# file = "C:\\Users\\karth\\Downloads\\Write_Data.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook.active #get active sheet from the excel if we have only one sheet
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value = "Welcome"
#
# workbook.save(file) # save the file after entering the data

# to write different data into the sheet
file = "C:\\Users\\karth\\Downloads\\Data_Sheet.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(1,1).value=123
sheet.cell(1,2).value="David"
sheet.cell(1,3).value="Engineer"

sheet.cell(2,1).value=456
sheet.cell(2,2).value="John"
sheet.cell(2,3).value="Developer"

sheet.cell(3,1).value=789
sheet.cell(3,2).value="Smith"
sheet.cell(3,3).value="Developer"

workbook.save(file) # save the file after entering the data
