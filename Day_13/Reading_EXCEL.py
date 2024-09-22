import openpyxl

# File --> Workbook --> Sheets --> Rows --> Cells

file = "C:/Users/karth/Downloads/Sheet1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
rows = sheet.max_row # count no.of rows in excel sheet
columns = sheet.max_column #count no.of columns in excel sheet

# Reading all the rows and columns from the excel sheet
for r in range(1, rows + 1):
    for c in range(1, columns + 1):
        print(sheet.cell(r, c).value, end = "   ")
    print()
